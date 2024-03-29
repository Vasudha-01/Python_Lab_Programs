#9A Download XKCD Comics 
import requests 
import os 
from bs4 import BeautifulSoup 
url = 'https://xkcd.com/1/' 
if not os.path.exists('xkcd_comics'): 
    os.makedirs('xkcd_comics') 
while True:  
    res = requests.get(url) 
    res.raise_for_status() 
    soup = BeautifulSoup(res.text, 'html.parser') 
    comic_elem = soup.select('#comic img') 
    if comic_elem == []: 
        print('Could not find comic image.') 
    else: 
        comic_url = 'https:' + comic_elem[0].get('src') 
        print(f'Downloading {comic_url}...') 
        res = requests.get(comic_url) 
        res.raise_for_status()  
        image_file = open(os.path.join('xkcd_comics', os.path.basename(comic_url)), 'wb') 
        for chunk in res.iter_content(100000): 
            image_file.write(chunk) 
        image_file.close() 
    prev_link = soup.select('a[rel="prev"]')[0] 
    if not prev_link: 
        break 
    url = 'https://xkcd.com' + prev_link.get('href') 
    print('All comics downloaded.') 



#9B Spreadsheet Operations 
from openpyxl import Workbook  
from openpyxl.styles import Font 
wb = Workbook()  
sheet = wb.active 
sheet.title = "Language" 
wb.create_sheet(title = "Capital") 
lang = ["Kannada", "Telugu", "Tamil"] 
state = ["Karnataka", "Telangana", "Tamil Nadu"] 
capital = ["Bengaluru", "Hyderabad", "Chennai"] 
code =['KA', 'TS', 'TN'] 
sheet.cell(row = 1, column = 1).value = "State" 
sheet.cell(row = 1, column = 2).value = "Language" 
sheet.cell(row = 1, column = 3).value = "Code" 
ft = Font(bold=True) 
for row in sheet["A1:C1"]: 
    for cell in row: 
        cell.font = ft 
for i in range(2,5): 
    sheet.cell(row = i, column = 1).value = state[i-2] 
    sheet.cell(row = i, column = 2).value = lang[i-2] 
    sheet.cell(row = i, column = 3).value = code[i-2] 
wb.save("demo.xlsx") 
sheet = wb["Capital"] 
sheet.cell(row = 1, column = 1).value = "State" 
sheet.cell(row = 1, column = 2).value = "Capital" 
sheet.cell(row = 1, column = 3).value = "Code" 
ft = Font(bold=True) 
for row in sheet["A1:C1"]: 
    for cell in row: 
        cell.font = ft 
for i in range(2,5): 
    sheet.cell(row = i, column = 1).value = state[i-2] 
    sheet.cell(row = i, column = 2).value = capital[i-2] 
    sheet.cell(row = i, column = 3).value = code[i-2] 
wb.save("demo.xlsx") 
srchCode = input("Enter state code for finding capital ") 
for i in range(2,5): 
    data = sheet.cell(row = i, column = 3).value 
    if data == srchCode: 
        print("Corresponding capital for code", srchCode, "is", sheet.cell(row = i, column = 2).value) 
sheet = wb["Language"] 
srchCode = input("Enter state code for finding language ") 
for i in range(2,5): 
    data = sheet.cell(row = i, column = 3).value 
    if data == srchCode: 
        print("Corresponding language for code", srchCode, "is", sheet.cell(row = i, column = 2).value) 
wb.close() 






